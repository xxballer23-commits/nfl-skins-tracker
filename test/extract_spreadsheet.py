"""Extract last season's numbers from the Excel tracker into a validation fixture.

The workbook stores only DERIVED values: a hand-entered 0/1 base skin per pick per
week, and a hand-entered aggregate bonus-skin total per team per week. It contains
no game scores, so the fixture synthesises scores that reproduce those derived
numbers, letting the real scoring engine be exercised end to end.

Usage: python3 test/extract_spreadsheet.py "../NFL Skins League (01.25.26)_v1.xlsx"
"""

import json
import sys
from pathlib import Path

import openpyxl

# Spreadsheet label -> pick id used by js/league.js
PICK_ID = {
    "Eagles W": "Eagles-W", "Giants L": "Giants-L", "49ers W": "49ers-W",
    "Patriots W": "Patriots-W", "Seahwks L": "Seahawks-L",
    "Bills W": "Bills-W", "Colts L": "Colts-L", "Panthers L": "Panthers-L",
    "Broncos W": "Broncos-W", "Jags L": "Jaguars-L",
    "Ravens W": "Ravens-W", "Texans W": "Texans-W", "Titans L": "Titans-L",
    "Cowboys L": "Cowboys-L", "Dolphins L": "Dolphins-L",
    "Chiefs W": "Chiefs-W", "Jets L": "Jets-L", "Bengals W": "Bengals-W",
    "Packers W": "Packers-W", "Falcons L": "Falcons-L",
    "Lions W": "Lions-W", "Browns L": "Browns-L", "Chargers W": "Chargers-W",
    "Raiders L": "Raiders-L", "Falcons W": "Falcons-W",
    "Commanders W": "Commanders-W", "Saints L": "Saints-L", "Bucs W": "Buccaneers-W",
    "Rams W": "Rams-W", "Cardinals L": "Cardinals-L",
}

TEAM_ID = {
    "Duriez / Cam": "duriez", "Alec / Nick": "alec", "Joe / Aidan": "joe",
    "Cummiskey / Culp": "cummiskey", "Little / Banks": "little",
    "Corley / Goddard": "corley",
}

# Weekly Scores: header row per team block, then 5 pick rows, skins, bonus, total.
BLOCK_START = {
    "Duriez / Cam": 5, "Alec / Nick": 15, "Joe / Aidan": 25,
    "Cummiskey / Culp": 35, "Little / Banks": 45, "Corley / Goddard": 55,
}

# column -> week id (Weekly Scores starts at week 2; week 1 lives on its own tab)
WEEK_COLS = {c: str(c - 1) for c in range(3, 20)}
WEEK_COLS.update({21: "WC", 22: "DIV", 23: "CC", 24: "SB"})
POSTSEASON = {"WC", "DIV", "CC", "SB"}

WEEK1_BLOCK_START = {
    "Duriez / Cam": 7, "Alec / Nick": 22, "Joe / Aidan": 30,
    "Cummiskey / Culp": 38, "Little / Banks": 46, "Corley / Goddard": 54,
}


def cell(ws, row, col):
    return ws.cell(row, col).value or 0


def synth_game(direction, earned_base, bonus_count):
    """Build a game result that yields exactly `earned_base` + `bonus_count` skins."""
    if not earned_base:
        # Played but did not earn: the opposite outcome, no bonus either way.
        return {"result": "L", "pointsFor": 17, "pointsAgainst": 20} if direction == "W" \
            else {"result": "W", "pointsFor": 20, "pointsAgainst": 17}
    if direction == "W":
        return [
            {"result": "W", "pointsFor": 20, "pointsAgainst": 17},  # no bonus
            {"result": "W", "pointsFor": 20, "pointsAgainst": 9},   # held under 9.5
            {"result": "W", "pointsFor": 42, "pointsAgainst": 6},   # both
        ][bonus_count]
    return [
        {"result": "L", "pointsFor": 17, "pointsAgainst": 20},
        {"result": "L", "pointsFor": 6, "pointsAgainst": 20},
        {"result": "L", "pointsFor": 3, "pointsAgainst": 45},
    ][bonus_count]


def main(xlsx_path, out_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Weekly Scores"]
    summary = wb["Summary"]
    perf = wb["Team Performance"]
    week1 = wb["Week 1"]

    results = {}
    problems = []

    for team_label, r0 in BLOCK_START.items():
        team_id = TEAM_ID[team_label]
        pick_rows = []
        for r in range(r0 + 1, r0 + 6):
            label = ws.cell(r, 2).value
            pick_rows.append((PICK_ID[label], label.rsplit(" ", 1)[1], r))

        for col, week_id in WEEK_COLS.items():
            bonus_left = int(cell(ws, r0 + 7, col))
            bases = {pid: int(cell(ws, r, col)) for pid, _, r in pick_rows}

            # Only picks that earned a base skin can carry bonuses (max 2 each).
            eligible = [(pid, d) for pid, d, _ in pick_rows if bases[pid]]
            capacity = 2 * len(eligible)
            if bonus_left > capacity:
                problems.append(
                    f"{team_label} {week_id}: {bonus_left} bonus skins but only "
                    f"{len(eligible)} scoring picks (max {capacity})"
                )

            assigned = {}
            for pid, _ in eligible:
                take = min(2, bonus_left)
                assigned[pid] = take
                bonus_left -= take

            for pid, direction, _ in pick_rows:
                if week_id in POSTSEASON and direction == "L":
                    # Deliberately encode a blowout loss: the engine must still return
                    # 0 because Lose picks are dead in the postseason.
                    entry = {"result": "L", "pointsFor": 3, "pointsAgainst": 45}
                else:
                    entry = synth_game(direction, bases[pid], assigned.get(pid, 0))
                results.setdefault(week_id, {})[pid] = entry

    # ---- expected values straight off the Summary tab ----
    expected_season = {}
    for r in range(25, 31):
        label = summary.cell(r, 3).value
        expected_season[TEAM_ID[label]] = {
            "base": int(summary.cell(r, 4).value),
            "bonus": int(summary.cell(r, 5).value),
            "total": int(summary.cell(r, 6).value),
        }

    matrix_teams = [TEAM_ID[summary.cell(32, c).value] for c in range(4, 10)]
    expected_by_week = {t: {} for t in matrix_teams}
    for r in range(33, 55):
        label = summary.cell(r, 3).value
        if label is None:
            break
        week_id = {"Wild Card": "WC", "Divisional": "DIV",
                   "Conf. Champ": "CC", "Super Bowl": "SB"}.get(label, str(label))
        for i, t in enumerate(matrix_teams):
            expected_by_week[t][week_id] = int(cell(summary, r, 4 + i))

    expected_picks = {}
    expected_draft = {}
    for r in range(6, 36):
        label = perf.cell(r, 5).value
        if label is None:
            break
        pick_id = PICK_ID[label]
        expected_picks[pick_id] = {
            "base": int(perf.cell(r, 6).value),
            "bonus": int(perf.cell(r, 7).value),
            "total": int(perf.cell(r, 8).value),
        }
        expected_draft[pick_id] = {
            "team": TEAM_ID[perf.cell(r, 2).value],
            "pickNo": int(perf.cell(r, 3).value),
            "round": int(str(perf.cell(r, 4).value).replace("Round ", "")),
        }

    # ---- Week 1 tab: base skins only, bonuses were switched off that week ----
    week1_results = {}
    expected_week1 = {}
    for team_label, r0 in WEEK1_BLOCK_START.items():
        team_id = TEAM_ID[team_label]
        for r in range(r0 + 1, r0 + 6):
            label = week1.cell(r, 2).value
            direction = label.rsplit(" ", 1)[1]
            earned = int(cell(week1, r, 3))
            week1_results[PICK_ID[label]] = synth_game(direction, earned, 0)
        expected_week1[team_id] = int(cell(week1, r0 + 6, 3))

    all_weeks = [str(w) for w in range(1, 19)] + ["WC", "DIV", "CC", "SB"]

    # Importable app state. Week 1 is left empty because last season's Summary
    # excluded it entirely; its numbers live on the separate Week 1 tab.
    app_state = {
        "version": 1,
        "settings": {
            "skinValue": summary.cell(14, 3).value,
            "bonusEnabledByWeek": {w: True for w in all_weeks},
        },
        "results": results,
    }

    fixture = {
        "source": Path(xlsx_path).name,
        "note": (
            "The workbook holds no game scores. Base skins are the workbook's own 0/1 "
            "values; bonus skins are synthesised from the workbook's per-team weekly "
            "bonus totals, spread across that team's scoring picks. Per-pick bonus "
            "attribution is therefore synthetic; per-team weekly and season totals are not."
        ),
        "problems": problems,
        "state": app_state,
        "expected": {
            "season": expected_season,
            "byWeek": expected_by_week,
            "picks": expected_picks,
            "draft": expected_draft,
            "mendoza": summary.cell(14, 7).value,
            "skinValue": summary.cell(14, 3).value,
        },
        "week1": {
            "results": week1_results,
            "expectedTotals": expected_week1,
            "mendoza": week1.cell(5, 8).value,
        },
    }

    Path(out_path).write_text(json.dumps(app_state, indent=1))
    fixture_path = Path("test/last-season.fixture.mjs")
    fixture_path.write_text(
        "// Generated by test/extract_spreadsheet.py. Do not edit by hand.\n"
        "export default " + json.dumps(fixture, indent=1) + ";\n"
    )
    print(f"wrote {out_path} and {fixture_path}")
    for p in problems:
        print("  DATA PROBLEM:", p)


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
